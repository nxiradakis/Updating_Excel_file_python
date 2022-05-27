import openpyxl
from openpyxl import load_workbook

workbook = load_workbook(filename="metrics.xlsx") #fetching the Excel file

#print(workbook.sheetnames)

sheet= workbook["Precision"] #assigning sheet title 
workbook.copy_worksheet(sheet) #making a copy of it
workbook.remove(sheet) #deleting existing sheet
sheet= workbook["Precision Copy"] #fetching/activating the buffered(copied) sheet
sheet.title = 'Precision' #renaming the copy to the original title
sheet.insert_rows(idx=2) #insert an new line in row no2
sheet["A2"] = precision1 #inserting the new values in this line
sheet["B2"] = precision2 #inserting the new values in this line


workbook.save(filename="metrics.xlsx") #saving the updated .xlsx file
